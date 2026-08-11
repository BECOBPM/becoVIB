import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
import io

# 1. 페이지 기본 설정
st.set_page_config(
    page_title="사업소별 설비 진동점검 관리 시스템",
    page_icon="⚙️",
    layout="wide"
)

# 2. 커스텀 CSS 스타일
st.markdown("""
<style>
    .main-title { font-size: 24px; font-weight: bold; color: #1F4E78; margin-bottom: 15px; }
    .site-badge { background-color: #EBF1F5; padding: 6px 14px; border-radius: 20px; font-weight: bold; color: #1F4E78; display: inline-block; margin-bottom: 15px; }
    .result-card { padding: 15px; border-radius: 8px; font-weight: bold; font-size: 18px; text-align: center; margin-top: 10px; }
</style>
""", unsafe_allow_html=True)

# 3. 사업소별 엑셀 표준 템플릿 생성 함수
def create_excel_template(site_name="공통"):
    wb = openpyxl.Workbook()
    
    # [시트 1] 진동측정기준
    ws_guide = wb.active
    ws_guide.title = "진동측정기준"
    ws_guide.views.sheetView[0].showGridLines = True
    
    ws_guide.merge_cells("A1:G1")
    ws_guide["A1"] = f"■ ISO 10816-3 진동 평가 기준표 ({site_name})"
    ws_guide["A1"].font = Font(name="맑은 고딕", size=14, bold=True, color="1F4E78")
    
    headers_guide = ["Machine Group", "설비 구분 / 용량", "기초 구분", "A (양호)", "B (장기운전)", "C (보수필요)", "D (즉시점검)"]
    ws_guide.append([])
    ws_guide.append(headers_guide)
    
    guide_data = [
        ["Group 1", "대형 전동기/설비 (P > 300 kW)", "강성 (Rigid)", "≤ 2.3 mm/s", "≤ 4.5 mm/s", "≤ 7.1 mm/s", "> 7.1 mm/s"],
        ["Group 1", "대형 전동기/설비 (P > 300 kW)", "유연 (Flexible)", "≤ 3.5 mm/s", "≤ 7.1 mm/s", "≤ 11.0 mm/s", "> 11.0 mm/s"],
        ["Group 2", "중형 전동기/설비 (15 kW < P ≤ 300 kW)", "강성 (Rigid)", "≤ 1.4 mm/s", "≤ 2.8 mm/s", "≤ 4.5 mm/s", "> 4.5 mm/s"],
        ["Group 2", "중형 전동기/설비 (15 kW < P ≤ 300 kW)", "유연 (Flexible)", "≤ 2.3 mm/s", "≤ 4.5 mm/s", "≤ 7.1 mm/s", "> 7.1 mm/s"],
    ]
    for row in guide_data:
        ws_guide.append(row)
        
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    for row in ws_guide.iter_rows(min_row=3, max_row=7, min_col=1, max_col=7):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if cell.row == 3:
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

    # [시트 2] 진동측정결과
    ws_data = wb.create_sheet(title="진동측정결과")
    ws_data.views.sheetView[0].showGridLines = True
    
    headers_data = ["사업소명", "측정일자", "설비명", "전동기(kW)", "측정위치", "방향", "진동속도 (rms)", "부하상태(%)", "판정 (A~D)", "이상 원인", "정비 우선순위"]
    ws_data.append(headers_data)
    
    for cell in ws_data[1]:
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    sample_rows = [
        [site_name, "2026-04-02", "FD Fan #1", 310, "1(전동기)", "X", 12.3, 30],
        [site_name, "2026-04-02", "FD Fan #1", 310, "1(전동기)", "Y", 2.8, 30],
        [site_name, "2026-04-02", "1차 냉각수 #1", 11, "2(펌프)", "Z", 4.0, 100],
        [site_name, "2026-04-02", "보일러 급수펌프 #1", 95, "1(전동기)", "X", 1.1, 100],
        [site_name, "2026-04-02", "보일러 급수펌프 #1", 95, "1(전동기)", "Y", 4.8, 100]
    ]
    
    for idx, r in enumerate(sample_rows, start=2):
        formula_grade = f'=IF(ISBLANK(G{idx}),"",IF(D{idx}>300,IF(G{idx}<=2.3,"A",IF(G{idx}<=4.5,"B",IF(G{idx}<=7.1,"C","D"))),IF(G{idx}<=1.4,"A",IF(G{idx}<=2.8,"B",IF(G{idx}<=4.5,"C","D")))))'
        formula_priority = f'=IF(I{idx}="A","양호",IF(I{idx}="B","양호",IF(I{idx}="C","보수필요",IF(I{idx}="D","즉시점검",""))))'
        row_data = r + [formula_grade, "", formula_priority]
        ws_data.append(row_data)

    fill_d = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_c = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_ab = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    ws_data.conditional_formatting.add("I2:I100", CellIsRule(operator='equal', formula=['"D"'], fill=fill_d))
    ws_data.conditional_formatting.add("I2:I100", CellIsRule(operator='equal', formula=['"C"'], fill=fill_c))
    ws_data.conditional_formatting.add("I2:I100", CellIsRule(operator='equal', formula=['"B"'], fill=fill_ab))
    ws_data.conditional_formatting.add("I2:I100", CellIsRule(operator='equal', formula=['"A"'], fill=fill_ab))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# 4. 최상위 사이드바 메뉴 (Navigation)
st.sidebar.title("📌 메뉴 (Navigation)")
main_menu = st.sidebar.radio(
    "구분 선택",
    ["🏢 사업소별 설비 관리", "📏 진동 측정 기준 (ISO 10816)"]
)

st.sidebar.divider()

# 5. 메인 메뉴 1: 사업소별 설비 관리
if main_menu == "🏢 사업소별 설비 관리":
    # [1단계] 사업소 선택
    selected_site = st.sidebar.selectbox(
        "🏢 사업소 선택",
        ["부산 사업소", "울산 사업소", "경주 사업소", "창원 사업소", "통합/전체"]
    )
    
    # [2단계] 선택된 사업소의 세부 메뉴 나열
    sub_menu = st.sidebar.radio(
        f"[{selected_site}] 세부 메뉴",
        ["📊 종합결과", "📁 점검결과 업로드", "📜 이력 관리"]
    )
    
    # 현재 선택된 사업소 태그 표시
    st.markdown(f"<div class='site-badge'>🏢 현재 선택된 사업소: <b>{selected_site}</b></div>", unsafe_allow_html=True)
    
    # [상세 1] 종합결과
    if sub_menu == "📊 종합결과":
        st.markdown(f"<div class='main-title'>📊 [{selected_site}] 설비 진동 점검 종합결과</div>", unsafe_allow_html=True)
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("사업소 점검 대상", "12 개")
        col2.metric("양호 (A/B)", "9 개", delta="정상")
        col3.metric("보수 필요 (C)", "2 개", delta="-1", delta_color="inverse")
        col4.metric("즉시 점검 (D)", "1 개", delta="+1", delta_color="inverse")
        
        st.divider()
        st.subheader(f"💡 [{selected_site}] 표준 엑셀 템플릿 다운로드")
        st.write(f"**{selected_site}** 명칭이 반영된 점검 양식 및 ISO 10816-3 기준표 시트가 첨부되어 있습니다.")
        excel_file = create_excel_template(selected_site)
        st.download_button(
            label=f"📥 {selected_site} 점검 엑셀 양식 다운로드 (.xlsx)",
            data=excel_file,
            file_name=f"{selected_site}_진동점검_표준양식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    # [상세 2] 점검결과 업로드
    elif sub_menu == "📁 점검결과 업로드":
        st.markdown(f"<div class='main-title'>📁 [{selected_site}] 점검결과 엑셀 업로드</div>", unsafe_allow_html=True)
        st.info(f"선택하신 **[{selected_site}]** 전용 진동점검 결과 엑셀 파일(.xlsx)을 업로드해주세요.")
        
        uploaded_file = st.file_uploader(f"[{selected_site}] 점검결과 파일 선택", type=["xlsx", "xls"])
        
        if uploaded_file is not None:
            try:
                df = pd.read_excel(uploaded_file, sheet_name="진동측정결과")
                st.success(f"[{selected_site}] 점검 데이터 로딩 완료!")
                st.dataframe(df, use_container_width=True)
            except Exception as e:
                st.error(f"파일을 읽는 도중 오류가 발생했습니다: {e}")

    # [상세 3] 이력 관리
    elif sub_menu == "📜 이력 관리":
        st.markdown(f"<div class='main-title'>📜 [{selected_site}] 설비별 진동점검 이력 관리</div>", unsafe_allow_html=True)
        st.write(f"**[{selected_site}]** 소속 회전기기(보일러 급수펌프, FD Fan, 송풍기 등)의 누적 측정 및 정비 이력 데이터입니다.")
        
        # 사업소별 샘플 이력 데이터
        history_df = pd.DataFrame({
            "측정일자": ["2026-04-01", "2026-03-15", "2026-02-10", "2026-01-20"],
            "사업소명": [selected_site] * 4,
            "설비명": ["FD Fan #1", "보일러 급수펌프 #1", "1차 냉각수펌프 #2", "FD Fan #1"],
            "측정위치": ["1(전동기) X", "1(전동기) Y", "2(펌프) Z", "1(전동기) X"],
            "진동속도(mm/s)": [12.3, 4.8, 1.2, 11.5],
            "판정": ["D (즉시점검)", "C (보수필요)", "A (양호)", "D (즉시점검)"],
            "조치사항 및 상태": ["베어링 수선 정비 예정", "구리스 정량 보충 완료", "정공정 이상없음", "진동 트렌드 지속 관찰 중"]
        })
        st.dataframe(history_df, use_container_width=True)

# 6. 메인 메뉴 2: 진동 측정 기준 (독립 페이지)
elif main_menu == "📏 진동 측정 기준 (ISO 10816)":
    st.markdown("<div class='main-title'>📏 ISO 10816-3 진동 평가 표준 기준 (공통)</div>", unsafe_allow_html=True)
    st.write("본 페이지는 전 사업소 공통으로 적용되는 **ISO 10816-3 표준 평가 기준** 및 **진동 계산기** 독립 안내 페이지입니다.")
    
    st.subheader("⚡ 실시간 진동 판정 계산기")
    col_a, col_b = st.columns(2)
    with col_a:
        kw_input = st.number_input("전동기 용량 (kW)", min_value=1.0, value=310.0, step=10.0)
    with col_b:
        rms_input = st.number_input("진동 속도 (mm/s rms)", min_value=0.0, value=12.3, step=0.1)
        
    if st.button("판정 계산하기"):
        if kw_input > 300:
            if rms_input <= 2.3: grade, color, text = "A", "#C6EFCE", "A (신품/매우양호)"
            elif rms_input <= 4.5: grade, color, text = "B", "#C6EFCE", "B (장기운전 가능)"
            elif rms_input <= 7.1: grade, color, text = "C", "#FFEB9C", "C (보수/계획정비 필요)"
            else: grade, color, text = "D", "#FFC7CE", "D (위험 / 즉시 점검)"
        else:
            if rms_input <= 1.4: grade, color, text = "A", "#C6EFCE", "A (신품/매우양호)"
            elif rms_input <= 2.8: grade, color, text = "B", "#C6EFCE", "B (장기운전 가능)"
            elif rms_input <= 4.5: grade, color, text = "C", "#FFEB9C", "C (보수/계획정비 필요)"
            else: grade, color, text = "D", "#FFC7CE", "D (위험 / 즉시 점검)"
            
        st.markdown(f"<div class='result-card' style='background-color:{color};'>판정 결과: {text}</div>", unsafe_allow_html=True)

    st.divider()
    st.subheader("📋 ISO 10816-3 기준표 (강성 기초 기준)")
    df_std = pd.DataFrame({
        "Machine Group": ["Group 2 (중형)", "Group 1 (대형)"],
        "전동기 용량": ["15 kW < P ≤ 300 kW", "300 kW < P ≤ 50 MW"],
        "A (양호)": ["≤ 1.4 mm/s", "≤ 2.3 mm/s"],
        "B (장기운전)": ["≤ 2.8 mm/s", "≤ 4.5 mm/s"],
        "C (보수필요)": ["≤ 4.5 mm/s", "≤ 7.1 mm/s"],
        "D (즉시점검)": ["> 4.5 mm/s", "> 7.1 mm/s"]
    })
    st.table(df_std)